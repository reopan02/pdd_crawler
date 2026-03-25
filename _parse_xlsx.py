import sys
import json

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

wb = openpyxl.load_workbook(
    r"C:\Users\huawei\Downloads\销售主题分析_渠道_20260324154901_149810759_1.xlsx",
    read_only=True,
)
ws = wb.active
headers = [c.value for c in ws[1]]
print("=== HEADERS ===")
for i, h in enumerate(headers):
    print(f"  Col {i}: {h}")

print("\n=== SAMPLE ROWS (2-5) ===")
for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=5, values_only=True), 2):
    print(f"Row {idx}:")
    for i, val in enumerate(row):
        print(f"  {headers[i]}: {val}")
    print()

all_rows = list(ws.iter_rows(values_only=True))
print(f"=== LAST ROW (row {len(all_rows)}) ===")
last = all_rows[-1]
for i, val in enumerate(last):
    print(f"  {headers[i]}: {val}")

print(f"\nTotal rows (incl header): {len(all_rows)}")
wb.close()
