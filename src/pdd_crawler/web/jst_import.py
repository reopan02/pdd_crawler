"""聚水潭 (JST) Excel sales data import engine."""

from __future__ import annotations

import io
import json
import re
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from typing import Any

from openpyxl import load_workbook

from pdd_crawler.web import db

MATCH_THRESHOLD = 0.75
AMBIGUOUS_GAP = 0.03
DETAIL_LIMIT = 200

_WEEKDAYS = ["一", "二", "三", "四", "五", "六", "日"]


def _weekday(d: date) -> str:
    return "周" + _WEEKDAYS[d.weekday()]


_EXCEL_FIELD_MAP = {
    "销售金额": "sales_amount",
    "退货金额": "refund_amount",
    "销售成本": "sales_cost",
    "退货成本": "refund_cost",
    "销售单数": "sales_order_count",
    "运费支出": "freight_expense",
}

_SHOP_COLUMN = "渠道"

_upload_cache: dict[str, dict[str, Any]] = {}
_preview_cache: dict[str, dict[str, Any]] = {}


def _normalise(name: str) -> str:
    s = str(name or "").strip().lower()
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("【", "[").replace("】", "]")
    s = s.replace("－", "-").replace("—", "-").replace("_", " ")
    s = re.sub(r"[|/\\,，;；]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def _compact(name: str) -> str:
    return re.sub(r"[\s\-\[\]\(\)]+", "", _normalise(name))


def _tokenise(name: str) -> set[str]:
    clean = re.sub(r"[\[\]()（）【】\-_/\\|,，;；]+", " ", _normalise(name))
    parts = [p for p in re.split(r"\s+", clean.strip()) if p]
    compact = _compact(name)
    tokens = set(parts)
    if compact:
        tokens.add(compact)
    return tokens


def _similarity(source: str, candidate: str) -> float:
    source_norm = _normalise(source)
    candidate_norm = _normalise(candidate)
    source_compact = _compact(source)
    candidate_compact = _compact(candidate)

    seq_ratio = SequenceMatcher(None, source_compact, candidate_compact).ratio()

    source_tokens = _tokenise(source_norm)
    candidate_tokens = _tokenise(candidate_norm)
    if source_tokens and candidate_tokens:
        token_ratio = len(source_tokens & candidate_tokens) / len(
            source_tokens | candidate_tokens
        )
    else:
        token_ratio = 0.0

    contain_ratio = 0.0
    if source_compact and candidate_compact:
        if source_compact in candidate_compact or candidate_compact in source_compact:
            contain_ratio = min(len(source_compact), len(candidate_compact)) / max(
                len(source_compact), len(candidate_compact)
            )

    return max(0.55 * seq_ratio + 0.3 * token_ratio + 0.15 * contain_ratio, seq_ratio)


def match_shop(
    source_name: str,
    db_shops: list[str],
) -> tuple[str | None, float, list[dict[str, Any]], bool]:
    """取相似度最高的已有店铺匹配；数据库中不存在的店铺返回未匹配。

    判断逻辑：取最高分店铺，若最高分 < 0.3 视为数据库中不存在该店铺。
    阈值极低，仅过滤完全无关的店铺名。
    """
    if not db_shops:
        return None, 0.0, [], False

    scored = [(shop, _similarity(source_name, shop)) for shop in db_shops]
    scored.sort(key=lambda item: item[1], reverse=True)

    top3 = [{"shop_name": shop, "score": round(score, 4)} for shop, score in scored[:3]]
    best_name, best_score = scored[0]

    # 最高分过低 → 数据库中不存在该店铺
    if best_score < 0.3:
        return None, best_score, top3, False

    return best_name, best_score, top3, False


def _safe_decimal(val: Any) -> Decimal:
    if val is None or val == "":
        return Decimal("0")
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"无法转换为数值: {val!r}") from exc


def _safe_int(val: Any) -> int:
    if val is None or val == "":
        return 0
    try:
        return int(float(str(val)))
    except (ValueError, TypeError) as exc:
        raise ValueError(f"无法转换为整数: {val!r}") from exc


def parse_xlsx(file_bytes: bytes, filename: str) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel 文件无有效工作表")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError("Excel 文件为空")

    headers = [str(h).strip() if h else "" for h in header_row]
    if _SHOP_COLUMN not in headers:
        raise ValueError(f"缺少必需列: {_SHOP_COLUMN}")

    missing_fields = [field for field in _EXCEL_FIELD_MAP if field not in headers]
    if missing_fields:
        raise ValueError(f"缺少必需列: {', '.join(missing_fields)}")

    shop_idx = headers.index(_SHOP_COLUMN)
    field_indices = {cn: headers.index(cn) for cn in _EXCEL_FIELD_MAP}

    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    total_rows = 0

    for row_num, row_vals in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row_vals is None:
            continue

        row_list = list(row_vals)
        shop_raw = row_list[shop_idx] if shop_idx < len(row_list) else None
        if shop_raw is None or str(shop_raw).strip() == "#":
            continue

        total_rows += 1
        source_shop = str(shop_raw).strip()

        try:
            parsed_row: dict[str, Any] = {"source_shop_name": source_shop}
            for cn, en in _EXCEL_FIELD_MAP.items():
                idx = field_indices[cn]
                val = row_list[idx] if idx < len(row_list) else None
                parsed_row[en] = (
                    _safe_int(val) if en == "sales_order_count" else _safe_decimal(val)
                )
            rows.append(parsed_row)
        except ValueError as exc:
            errors.append({"row": row_num, "shop": source_shop, "error": str(exc)})

    wb.close()
    return {
        "rows": rows,
        "errors": errors,
        "total_rows": total_rows,
        "filename": filename,
    }


def get_uploaded_payload(upload_token: str) -> dict[str, Any] | None:
    return _upload_cache.get(upload_token)


def _jsonable_errors(stage: str, errors: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{"stage": stage, **item} for item in errors]


def _decode_json_field(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value


async def register_upload(
    upload_token: str, filename: str, parsed: dict[str, Any]
) -> dict[str, Any]:
    uploaded_at = datetime.now(timezone.utc).isoformat()
    payload = {
        "upload_token": upload_token,
        "filename": filename,
        "parsed": parsed,
        "uploaded_at": uploaded_at,
    }
    _upload_cache[upload_token] = payload
    _preview_cache.pop(upload_token, None)

    log_id = str(uuid.uuid4())
    parse_errors = parsed["errors"]
    async with db.get_conn() as conn:
        await conn.execute(
            """
            INSERT INTO import_logs
                (id, file_name, upload_token, total_rows, parse_error_count, status, error_details)
            VALUES ($1,$2,$3,$4,$5,'uploaded',$6)
            ON CONFLICT (upload_token) DO UPDATE SET
                file_name = EXCLUDED.file_name,
                total_rows = EXCLUDED.total_rows,
                parse_error_count = EXCLUDED.parse_error_count,
                status = 'uploaded',
                preview_time = NULL,
                commit_time = NULL,
                latest_preview_id = NULL,
                committed_preview_id = NULL,
                preview_biz_date = NULL,
                matched_count = 0,
                unmatched_count = 0,
                ambiguous_count = 0,
                to_insert_count = 0,
                duplicate_count = 0,
                inserted_count = 0,
                skipped_count = 0,
                failed_count = 0,
                error_details = EXCLUDED.error_details
            """,
            log_id,
            filename,
            upload_token,
            parsed["total_rows"],
            len(parse_errors),
            json.dumps(_jsonable_errors("parse", parse_errors), ensure_ascii=False),
        )
    return payload


async def get_db_shops() -> list[str]:
    async with db.get_conn() as conn:
        records = await conn.fetch(
            "SELECT DISTINCT shop_name FROM shop_daily_reports ORDER BY shop_name"
        )
        return [record["shop_name"] for record in records]


async def _get_log_by_upload_token(
    conn: Any, upload_token: str
) -> dict[str, Any] | None:
    row = await conn.fetchrow(
        "SELECT * FROM import_logs WHERE upload_token = $1",
        upload_token,
    )
    return dict(row) if row else None


async def build_preview(
    upload_token: str,
    parsed: dict[str, Any],
    filename: str,
    biz_date: date,
) -> dict[str, Any]:
    db_shops = await get_db_shops()
    rows = parsed["rows"]
    parse_errors = parsed["errors"]

    # Phase 1: 对每行计算最佳匹配
    candidates: list[dict[str, Any]] = []
    for row in rows:
        source = row["source_shop_name"]
        matched_name, score, top3, _ = match_shop(source, db_shops)
        candidates.append(
            {
                "row": row,
                "source": source,
                "matched_name": matched_name,
                "score": score,
                "top3": top3,
            }
        )

    # Phase 2: 每个 DB 店铺只分配给置信度最高的 Excel 行
    # 按 matched_name 分组，取 score 最高的
    best_for_shop: dict[str, int] = {}  # db_shop_name → index in candidates
    for i, c in enumerate(candidates):
        if c["matched_name"] is None:
            continue
        shop = c["matched_name"]
        if (
            shop not in best_for_shop
            or c["score"] > candidates[best_for_shop[shop]]["score"]
        ):
            best_for_shop[shop] = i

    # 被选中的 index 集合
    winner_indices = set(best_for_shop.values())

    matched_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []
    match_details: list[dict[str, Any]] = []

    for i, c in enumerate(candidates):
        source = c["source"]
        top3 = c["top3"]

        if c["matched_name"] is not None and i in winner_indices:
            # 匹配成功
            matched_rows.append({**c["row"], "shop_name_matched": c["matched_name"]})
            match_details.append(
                {
                    "source_shop_name": source,
                    "matched_shop_name": c["matched_name"],
                    "score": round(c["score"], 4),
                    "is_ambiguous": False,
                    "top_candidates": top3,
                    "status": "matched",
                }
            )
        else:
            # 未匹配：DB 无此店铺，或该 DB 店铺已被更高分的行占用
            if c["matched_name"] is None:
                reason = "数据库中不存在该店铺"
            else:
                winner_idx = best_for_shop.get(c["matched_name"])
                winner_source = (
                    candidates[winner_idx]["source"] if winner_idx is not None else "?"
                )
                reason = f"店铺已被更高置信度的行匹配（{winner_source}）"
            unmatched_rows.append(
                {
                    "source_shop_name": source,
                    "reason": reason,
                    "top_candidates": top3,
                }
            )
            match_details.append(
                {
                    "source_shop_name": source,
                    "matched_shop_name": None,
                    "score": round(c["score"], 4),
                    "is_ambiguous": False,
                    "top_candidates": top3,
                    "status": "unmatched",
                    "reason": reason,
                }
            )

    to_insert: list[dict[str, Any]] = []
    duplicate_count = 0

    async with db.get_conn() as conn:
        existing = await conn.fetch(
            "SELECT biz_date, shop_name, source_file_name "
            "FROM jst_sales_imports "
            "WHERE biz_date = $1 AND source_file_name = $2",
            biz_date,
            filename,
        )
        existing_keys = {
            (record["biz_date"], record["shop_name"], record["source_file_name"])
            for record in existing
        }

        for row in matched_rows:
            row_key = (biz_date, row["shop_name_matched"], filename)
            if row_key in existing_keys:
                duplicate_count += 1
            else:
                to_insert.append(row)

        preview_id = str(uuid.uuid4())
        created_at = datetime.now(timezone.utc).isoformat()
        snapshot = {
            "preview_id": preview_id,
            "upload_token": upload_token,
            "filename": filename,
            "biz_date": biz_date.isoformat(),
            "to_insert": to_insert,
            "match_details": match_details,
            "unmatched_rows": unmatched_rows,
            "parse_errors": parse_errors,
            "created_at": created_at,
            "stats": {
                "total_rows": parsed["total_rows"],
                "matched_count": len(matched_rows),
                "unmatched_count": len(unmatched_rows),
                "ambiguous_count": sum(
                    1 for detail in match_details if detail.get("is_ambiguous")
                ),
                "to_insert_count": len(to_insert),
                "duplicate_count": duplicate_count,
                "parse_error_count": len(parse_errors),
            },
        }
        _preview_cache[upload_token] = snapshot

        log_id = str(uuid.uuid4())
        parse_error_payload = _jsonable_errors("parse", parse_errors)
        await conn.execute(
            """
            INSERT INTO import_logs
                (id, file_name, upload_token, preview_time, total_rows,
                 matched_count, unmatched_count, ambiguous_count, to_insert_count,
                 duplicate_count, parse_error_count, latest_preview_id, preview_biz_date,
                 status, error_details)
            VALUES ($1,$2,$3,now(),$4,$5,$6,$7,$8,$9,$10,$11,$12,'previewed',$13)
            ON CONFLICT (upload_token) DO UPDATE SET
                file_name = EXCLUDED.file_name,
                preview_time = now(),
                total_rows = EXCLUDED.total_rows,
                matched_count = EXCLUDED.matched_count,
                unmatched_count = EXCLUDED.unmatched_count,
                ambiguous_count = EXCLUDED.ambiguous_count,
                to_insert_count = EXCLUDED.to_insert_count,
                duplicate_count = EXCLUDED.duplicate_count,
                parse_error_count = EXCLUDED.parse_error_count,
                latest_preview_id = EXCLUDED.latest_preview_id,
                preview_biz_date = EXCLUDED.preview_biz_date,
                status = 'previewed',
                inserted_count = 0,
                skipped_count = 0,
                failed_count = 0,
                committed_preview_id = NULL,
                commit_time = NULL,
                error_details = EXCLUDED.error_details
            """,
            log_id,
            filename,
            upload_token,
            parsed["total_rows"],
            snapshot["stats"]["matched_count"],
            snapshot["stats"]["unmatched_count"],
            snapshot["stats"]["ambiguous_count"],
            snapshot["stats"]["to_insert_count"],
            snapshot["stats"]["duplicate_count"],
            snapshot["stats"]["parse_error_count"],
            preview_id,
            biz_date,
            json.dumps(parse_error_payload, ensure_ascii=False),
        )

        log_row = await _get_log_by_upload_token(conn, upload_token)
        if not log_row:
            raise ValueError("导入日志创建失败")

        await conn.execute(
            "DELETE FROM import_unmatched WHERE import_log_id = $1",
            log_row["id"],
        )
        for unmatched in unmatched_rows:
            await conn.execute(
                """
                INSERT INTO import_unmatched
                    (id, import_log_id, source_shop_name, reason, top_candidates)
                VALUES ($1,$2,$3,$4,$5)
                """,
                str(uuid.uuid4()),
                log_row["id"],
                unmatched["source_shop_name"],
                unmatched["reason"],
                json.dumps(unmatched["top_candidates"], ensure_ascii=False),
            )

    return {
        "preview_id": preview_id,
        "upload_token": upload_token,
        "filename": filename,
        "biz_date": biz_date.isoformat(),
        "created_at": created_at,
        "log_id": log_row["id"],
        "stats": snapshot["stats"],
        "match_details": match_details[:DETAIL_LIMIT],
        "unmatched_rows": unmatched_rows[:DETAIL_LIMIT],
        "parse_errors": parse_errors[:DETAIL_LIMIT],
    }


async def commit_import(upload_token: str, preview_id: str) -> dict[str, Any]:
    async with db.get_conn() as conn:
        log = await _get_log_by_upload_token(conn, upload_token)
        if not log:
            raise ValueError("导入日志不存在，请重新上传")

        if log.get("status") == "committed":
            if log.get("committed_preview_id") == preview_id:
                return {
                    "status": "already_committed",
                    "inserted_count": log["inserted_count"],
                    "skipped_count": log["skipped_count"],
                    "failed_count": log["failed_count"],
                    "log_id": log["id"],
                }
            raise ValueError("当前批次已用其他预览快照提交，请重新上传文件")

        if log.get("latest_preview_id") != preview_id:
            raise ValueError("仅允许提交最近一次预览快照，请重新预览后再写入")

        snapshot = _preview_cache.get(upload_token)
        if not snapshot:
            raise ValueError("预览数据已过期，请重新上传并预览")
        if snapshot["preview_id"] != preview_id:
            raise ValueError("预览快照已失效，请重新预览")

        to_insert = snapshot["to_insert"]
        filename = snapshot["filename"]
        biz_date = date.fromisoformat(snapshot["biz_date"])
        existing_errors = _decode_json_field(log.get("error_details")) or []

        inserted = 0
        skipped = 0
        failed = 0
        commit_errors: list[dict[str, Any]] = []

        async with conn.transaction():
            for row in to_insert:
                try:
                    result = await conn.execute(
                        """
                        INSERT INTO jst_sales_imports
                            (id, biz_date, shop_name, shop_name_matched,
                             source_shop_name, sales_amount, refund_amount,
                             sales_cost, refund_cost, sales_order_count,
                             freight_expense, source_file_name)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
                        ON CONFLICT (biz_date, shop_name, source_file_name) DO NOTHING
                        """,
                        str(uuid.uuid4()),
                        biz_date,
                        row["shop_name_matched"],
                        row["shop_name_matched"],
                        row["source_shop_name"],
                        float(row["sales_amount"]),
                        float(row["refund_amount"]),
                        float(row["sales_cost"]),
                        float(row["refund_cost"]),
                        row["sales_order_count"],
                        float(row["freight_expense"]),
                        filename,
                    )
                    if result == "INSERT 0 1":
                        inserted += 1
                    else:
                        skipped += 1
                except Exception as exc:
                    failed += 1
                    commit_errors.append(
                        {
                            "shop": row["source_shop_name"],
                            "matched_shop_name": row["shop_name_matched"],
                            "error": str(exc),
                        }
                    )

            all_errors = list(existing_errors) + _jsonable_errors(
                "commit", commit_errors
            )
            await conn.execute(
                """
                UPDATE import_logs SET
                    commit_time = now(),
                    inserted_count = $2,
                    skipped_count = $3,
                    failed_count = $4,
                    status = 'committed',
                    committed_preview_id = $5,
                    error_details = $6::jsonb
                WHERE upload_token = $1
                """,
                upload_token,
                inserted,
                skipped,
                failed,
                preview_id,
                json.dumps(all_errors, ensure_ascii=False),
            )

            # Upsert rows into shop_daily_reports so /data preview
            # reflects newly imported JST data immediately.
            for row in to_insert:
                await conn.execute(
                    """
                    INSERT INTO shop_daily_reports
                        (id, shop_name, data_date, weekday,
                         sales_amount, refund_amount, sales_cost,
                         refund_cost, sales_order_count, freight_expense)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                    ON CONFLICT (shop_name, data_date) DO UPDATE SET
                        sales_amount      = EXCLUDED.sales_amount,
                        refund_amount     = EXCLUDED.refund_amount,
                        sales_cost        = EXCLUDED.sales_cost,
                        refund_cost       = EXCLUDED.refund_cost,
                        sales_order_count = EXCLUDED.sales_order_count,
                        freight_expense   = EXCLUDED.freight_expense
                    """,
                    str(uuid.uuid4()),
                    row["shop_name_matched"],
                    biz_date,
                    _weekday(biz_date),
                    float(row["sales_amount"]),
                    float(row["refund_amount"]),
                    float(row["sales_cost"]),
                    float(row["refund_cost"]),
                    row["sales_order_count"],
                    float(row["freight_expense"]),
                )

    _preview_cache.pop(upload_token, None)
    _upload_cache.pop(upload_token, None)

    return {
        "status": "committed",
        "inserted_count": inserted,
        "skipped_count": skipped,
        "failed_count": failed,
        "errors": commit_errors[:DETAIL_LIMIT],
        "log_id": log["id"],
    }


async def _hydrate_import_log(row: Any) -> dict[str, Any] | None:
    if not row:
        return None
    log = dict(row)
    for key in ("upload_time", "preview_time", "commit_time", "preview_biz_date"):
        if log.get(key) and hasattr(log[key], "isoformat"):
            log[key] = log[key].isoformat()
    log["error_details"] = _decode_json_field(log.get("error_details")) or []
    return log


async def get_import_log(log_id: str) -> dict[str, Any] | None:
    async with db.get_conn() as conn:
        log_row = await conn.fetchrow("SELECT * FROM import_logs WHERE id = $1", log_id)
        log = await _hydrate_import_log(log_row)
        if not log:
            return None
        unmatched_rows = await conn.fetch(
            """
            SELECT source_shop_name, reason, top_candidates
            FROM import_unmatched
            WHERE import_log_id = $1
            ORDER BY created_at, source_shop_name
            """,
            log_id,
        )
        log["unmatched_rows"] = [
            {
                "source_shop_name": row["source_shop_name"],
                "reason": row["reason"],
                "top_candidates": _decode_json_field(row["top_candidates"]) or [],
            }
            for row in unmatched_rows
        ]
        return log


async def get_import_log_by_token(upload_token: str) -> dict[str, Any] | None:
    async with db.get_conn() as conn:
        row = await conn.fetchrow(
            "SELECT * FROM import_logs WHERE upload_token = $1",
            upload_token,
        )
        log = await _hydrate_import_log(row)
        if not log:
            return None
        unmatched_rows = await conn.fetch(
            """
            SELECT source_shop_name, reason, top_candidates
            FROM import_unmatched
            WHERE import_log_id = $1
            ORDER BY created_at, source_shop_name
            """,
            log["id"],
        )
        log["unmatched_rows"] = [
            {
                "source_shop_name": row["source_shop_name"],
                "reason": row["reason"],
                "top_candidates": _decode_json_field(row["top_candidates"]) or [],
            }
            for row in unmatched_rows
        ]
        return log
