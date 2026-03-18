"""Excel export with styled formatting.

Generates a multi-sheet XLSX workbook (one sheet per shop) with:
- Title row with shop name and month
- Grouped column headers (店铺基础数据 / 付费推广数据 / 营业费用)
- Styled data rows with number formatting
- Summary (合计) row
- Frozen panes (top 3 rows + left 2 columns)

Mirrors the export project's exceljs output using openpyxl.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pdd_crawler.web import data_store

# ── Column definitions ───────────────────────────────────

COLUMNS = [
    {"key": "date", "label": "日期", "group": "", "width": 14},
    {"key": "weekday", "label": "星期", "group": "", "width": 10},
    {
        "key": "payment_amount",
        "label": "支付金额",
        "group": "店铺基础数据",
        "width": 14,
    },
    {
        "key": "promotion_cost",
        "label": "全站推广",
        "group": "付费推广数据",
        "width": 14,
    },
    {
        "key": "marketing_cost",
        "label": "评价有礼+跨店满减",
        "group": "营业费用",
        "width": 22,
    },
    {"key": "after_sale_cost", "label": "售后费用", "group": "营业费用", "width": 14},
    {
        "key": "tech_service_fee",
        "label": "技术服务费",
        "group": "营业费用",
        "width": 14,
    },
    {"key": "other_cost", "label": "其他费用", "group": "营业费用", "width": 14},
    {"key": "platform_refund", "label": "平台返还", "group": "营业费用", "width": 14},
]

# ── Styles ───────────────────────────────────────────────

_thin = Side(style="thin", color="D0D0D0")
_all_borders = Border(top=_thin, left=_thin, bottom=_thin, right=_thin)
_medium_top = Border(
    top=Side(style="medium", color="409EFF"),
    left=_thin,
    bottom=_thin,
    right=_thin,
)

_header_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
_title_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
_summary_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

_group_fills: dict[str, PatternFill] = {
    "店铺基础数据": PatternFill(
        start_color="E8F4FD", end_color="E8F4FD", fill_type="solid"
    ),
    "付费推广数据": PatternFill(
        start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
    ),
    "营业费用": PatternFill(
        start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"
    ),
}

_title_font = Font(bold=True, size=14, color="FFFFFF")
_group_font = Font(bold=True, size=11, color="303133")
_header_font = Font(bold=True, size=10, color="606266")
_data_font = Font(size=10)
_nodata_font = Font(size=10, color="CCCCCC")
_summary_font = Font(bold=True, size=10)
_center = Alignment(horizontal="center", vertical="center")
_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
_right = Alignment(horizontal="right", vertical="center")
_number_fmt = "#,##0.00"


# ── Group span calculation ───────────────────────────────


def _compute_group_spans() -> list[dict[str, Any]]:
    """Compute merged-cell spans for the group header row."""
    spans: list[dict[str, Any]] = []
    current_group = ""
    start_idx = 0

    for i, col in enumerate(COLUMNS):
        g = col["group"]
        if g != current_group:
            if current_group and start_idx < i:
                spans.append({"group": current_group, "start": start_idx + 1, "end": i})
            current_group = g
            start_idx = i

    # Close last group
    if current_group:
        spans.append(
            {"group": current_group, "start": start_idx + 1, "end": len(COLUMNS)}
        )

    return spans


# ── Main export function ─────────────────────────────────


def generate_xlsx(shop_names: list[str], month: str) -> bytes:
    """Generate a styled XLSX workbook with one sheet per shop.

    Args:
        shop_names: List of shop names to include.
        month: YYYY-MM string.

    Returns:
        XLSX file content as bytes.
    """
    wb = Workbook()
    # Remove the default sheet
    if wb.active:
        wb.remove(wb.active)

    col_count = len(COLUMNS)
    group_spans = _compute_group_spans()

    try:
        mon_num = int(month.split("-")[1])
    except (IndexError, ValueError):
        mon_num = 0

    for shop_name in shop_names:
        template = data_store.build_template_data(shop_name, month)
        sheet_name = shop_name[:31] if len(shop_name) > 31 else shop_name
        ws = wb.create_sheet(title=sheet_name)

        # ── Row 1: Title ─────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        title_cell = ws.cell(
            row=1, column=1, value=f"{mon_num}月份 {shop_name} 全店数据"
        )
        title_cell.font = _title_font
        title_cell.fill = _title_fill
        title_cell.alignment = _center
        ws.row_dimensions[1].height = 32

        # ── Row 2: Group headers ─────────────────────────
        ws.row_dimensions[2].height = 24
        # Fill all cells in row 2 with header background first
        for ci in range(1, col_count + 1):
            cell = ws.cell(row=2, column=ci)
            cell.fill = _header_fill
            cell.border = _all_borders

        for span in group_spans:
            s, e = span["start"], span["end"]
            if s < e:
                ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
            cell = ws.cell(row=2, column=s, value=span["group"])
            cell.font = _group_font
            cell.alignment = _center
            cell.fill = _group_fills.get(span["group"], _header_fill)
            cell.border = _all_borders

        # ── Row 3: Column headers ────────────────────────
        ws.row_dimensions[3].height = 28
        for ci, col_def in enumerate(COLUMNS, 1):
            cell = ws.cell(row=3, column=ci, value=col_def["label"])
            cell.font = _header_font
            cell.fill = _header_fill
            cell.alignment = _center_wrap
            cell.border = _all_borders

        # ── Data rows ────────────────────────────────────
        for ri, row_data in enumerate(template["rows"], 4):
            for ci, col_def in enumerate(COLUMNS, 1):
                key = col_def["key"]
                if key in ("date", "weekday"):
                    val = row_data.get(key, "")
                elif key == "tech_service_fee":
                    val = abs(row_data.get(key, 0)) if row_data["has_data"] else ""
                else:
                    val = row_data.get(key, 0) if row_data["has_data"] else ""

                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = _all_borders

                if ci <= 2:
                    cell.alignment = _center
                    cell.font = _nodata_font if not row_data["has_data"] else _data_font
                else:
                    cell.alignment = _right
                    cell.font = _nodata_font if not row_data["has_data"] else _data_font
                    if isinstance(val, (int, float)) and val != 0:
                        cell.number_format = _number_fmt

        # ── Summary row ──────────────────────────────────
        summary_row_idx = 4 + len(template["rows"])
        totals = template["totals"]

        for ci, col_def in enumerate(COLUMNS, 1):
            key = col_def["key"]
            if key == "date":
                val: Any = "合计"
            elif key == "weekday":
                val = ""
            elif key == "tech_service_fee":
                val = abs(totals.get(key, 0))
            else:
                val = totals.get(key, 0)

            cell = ws.cell(row=summary_row_idx, column=ci, value=val)
            cell.font = _summary_font
            cell.fill = _summary_fill
            cell.border = _medium_top

            if ci <= 2:
                cell.alignment = _center
            else:
                cell.alignment = _right
                if isinstance(val, (int, float)) and val != 0:
                    cell.number_format = _number_fmt

        # ── Column widths ────────────────────────────────
        for ci, col_def in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_def["width"]

        # ── Freeze panes (top 3 rows + left 2 columns) ──
        ws.freeze_panes = "C4"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
