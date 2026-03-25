"""Data management API endpoints.

Provides CRUD + query + export for shop daily report data.
"""

from __future__ import annotations

import io
import json

import asyncpg
from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from pdd_crawler.web import data_store as _ds

router = APIRouter(prefix="/data", tags=["data"])


# ── Meta endpoints ────────────────────────────────────────


@router.get("/shops")
async def get_shops():
    """Return all distinct shop names."""
    return {"shops": await _ds.store.get_shops()}


@router.get("/months")
async def get_months():
    """Return all distinct YYYY-MM months, newest first."""
    return {"months": await _ds.store.get_months()}


# ── Query endpoint ────────────────────────────────────────


@router.get("/query")
async def query_data(request: Request):
    """Query rows filtered by shops and/or month.

    Query params:
        shops: comma-separated shop names (optional)
        month: YYYY-MM string (optional)
    """
    params = request.query_params
    shops_param = params.get("shops", "")
    shops = [s for s in shops_param.split(",") if s] if shops_param else None
    month = params.get("month", "") or None
    rows = await _ds.store.query(shops, month)
    return {"rows": rows, "count": len(rows)}


# ── CRUD endpoints ────────────────────────────────────────


@router.post("/rows")
async def add_row(request: Request):
    """Add a new data row (upsert on shop_name + data_date conflict)."""
    body = await request.json()
    if not body.get("shop_name") or not body.get("data_date"):
        raise HTTPException(status_code=422, detail="shop_name 和 data_date 为必填项")
    try:
        row = await _ds.store.add_row(body)
    except asyncpg.UniqueViolationError:
        raise HTTPException(status_code=409, detail="该店铺该日期的数据已存在")
    return {"success": True, "row": row}


@router.put("/rows/{row_id}")
async def update_row(row_id: str, request: Request):
    """Update numeric fields of an existing row."""
    body = await request.json()
    row = await _ds.store.update_row(row_id, body)
    if row is None:
        raise HTTPException(status_code=404, detail="数据行不存在")
    return {"success": True, "row": row}


@router.delete("/rows/{row_id}")
async def delete_row(row_id: str):
    """Delete a data row by id."""
    ok = await _ds.store.delete_row(row_id)
    if not ok:
        raise HTTPException(status_code=404, detail="数据行不存在")
    return {"success": True}


# ── Upload endpoint ───────────────────────────────────────


@router.post("/upload")
async def upload_data(files: list[UploadFile] = File(...)):
    """Import data rows from JSON file(s)."""
    total = 0
    for f in files:
        content = await f.read()
        try:
            text = content.decode("utf-8")
        except Exception:
            try:
                text = content.decode("gb18030")
            except Exception:
                continue
        total += await _ds.store.import_from_json_file(text)
    return {"success": True, "count": total}


# ── Export endpoint ───────────────────────────────────────

_COLUMNS = [
    ("data_date", "日期"),
    ("weekday", "星期"),
    ("payment_amount", "支付金额"),
    ("promotion_cost", "全站推广"),
    ("marketing_cost", "评价有礼+跨店满减"),
    ("after_sale_cost", "售后费用"),
    ("tech_service_fee", "技术服务费"),
    ("other_cost", "其他费用"),
    ("platform_refund", "平台返还"),
    ("sales_amount", "销售金额"),
    ("refund_amount", "退货金额"),
    ("sales_cost", "销售成本"),
    ("refund_cost", "退货成本"),
    ("sales_order_count", "销售单数"),
    ("freight_expense", "运费支出"),
]


@router.get("/export")
async def export_xlsx(request: Request):
    """Export filtered data as Excel (.xlsx)."""
    params = request.query_params
    shops_param = params.get("shops", "")
    shops = [s for s in shops_param.split(",") if s] if shops_param else None
    month = params.get("month", "") or None

    rows = await _ds.store.query(shops, month)

    wb = Workbook()
    ws = wb.active
    ws.title = month or "数据"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")

    # Group header row
    group_headers = [
        ("", 1),
        ("", 1),
        ("店铺基础数据", 1),
        ("付费推广数据", 1),
        ("营业费用", 5),
        ("聚水潭数据", 6),
    ]
    col = 1
    for label, span in group_headers:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        if span > 1:
            ws.merge_cells(
                start_row=1, start_column=col, end_row=1, end_column=col + span - 1
            )
        col += span

    # Column header row
    for ci, (_, label) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Shop name column before data columns (prepend)
    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="店铺名称").font = header_font
    ws.cell(row=2, column=1, value="店铺名称").font = header_font

    # Data rows
    for ri, row in enumerate(rows, start=3):
        ws.cell(row=ri, column=1, value=row.get("shop_name", ""))
        for ci, (key, _) in enumerate(_COLUMNS, start=2):
            ws.cell(row=ri, column=ci, value=row.get(key, ""))

    # Auto column width
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"店铺数据_{month or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
